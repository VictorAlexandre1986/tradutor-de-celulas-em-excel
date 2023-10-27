import openpyxl
from translate import Translator

#Carregar a planilha excel
# workbook = openpyxl.load_workbook('Documents/Planilha.xlsx')
workbook = openpyxl.load_workbook('Planilha.xlsx')

#Define a planilha ativa
sheet = workbook.active

#Instanciar o tradutor
translator = Translator(from_lang="english", to_lang="portuguese")

#Define a função para traduzir o conteúdo de uma célula
def translate_cell(cell):
    translated_text = translator.translate(cell)  
    return translated_text

# Itera pelas células na planilha e traduz
for row in sheet.iter_rows():
    for cell in row:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = translate_cell(cell.value)
            
#Salva a planilha traduzida
# workbook.save('Documents/Planilha Traduzida')
workbook.save('Planilha Traduzida.xlsx')
workbook.close